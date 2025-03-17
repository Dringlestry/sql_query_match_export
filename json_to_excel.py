import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 读取 JSON 文件
with open('output2.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

# 用于存储最终数据的列表
rows = []

# 遍历每个顶级键值对
for key, value in data.items():
    fields = {
        'native_gc': value.get('native_gc', []),
        'native_hj': value.get('native_hj', []),
        'native_kf': value.get('native_kf', []),
        'native_zx': value.get('native_zx', []),
        'complaints_ticket': value.get('complaints_ticket', []),
        'matters_ticket': value.get('matters_ticket', []),
        'repairs_ticket': value.get('repairs_ticket', [])
    }

    for field_name, field_data in fields.items():
        for item in field_data:
            for query_item in item.get('query', []):
                for query_key, query_value in query_item.items():
                    # 合并 query_key 和 query_value
                    query_combined = f"{query_key}: {query_value}"
                    row = [key, field_name, item.get('hot', ''), item.get('ids', ''), query_combined]   # 一条一条的往excel插入数据
                    rows.append(row)

# 创建 DataFrame
df = pd.DataFrame(rows, columns=['顶级键', 'field_name', 'hot', 'ids', 'query_combined'])

# 将 DataFrame 保存到 Excel 文件
df.to_excel('output12.xlsx', index=False)

# 加载生成的 Excel 文件
wb = load_workbook('output12.xlsx')
ws = wb.active

# 遍历每一列
for col_idx in range(1, ws.max_column + 1):
    start_row = 2  # 从第二行开始（跳过表头）
    current_value = ws.cell(row=start_row, column=col_idx).value
    for row_idx in range(3, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value == current_value:
            continue
        else:
            if row_idx - start_row > 1:
                # 合并单元格
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{row_idx - 1}')
                # 设置合并后单元格的对齐方式为居中
                merged_cell = ws.cell(row=start_row, column=col_idx)
                merged_cell.alignment = Alignment(vertical='center', horizontal='center')
            start_row = row_idx
            current_value = cell_value
    # 处理最后一组连续相同的单元格
    if ws.max_row - start_row > 0:
        col_letter = get_column_letter(col_idx)
        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{ws.max_row}')
        merged_cell = ws.cell(row=start_row, column=col_idx)
        merged_cell.alignment = Alignment(vertical='center', horizontal='center')

# 保存修改后的 Excel 文件
wb.save('output12.xlsx')